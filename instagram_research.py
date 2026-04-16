"""
Bridge and Tunnel Brewery - Instagram Research Tool
----------------------------------------------------
Uses Social Blade to look up Instagram follower counts and stats
for bands and artists. No login required.

Usage:
  python instagram_research.py --handle "bandname"
  python instagram_research.py --handle "band1,band2,band3"
  python instagram_research.py --scrape-brewery
  python instagram_research.py --file "bands.txt"

Results are saved to:
  C:\Users\ADmin2\Desktop\BT_Talent_Database.xlsx
"""

import sys
import os
import time
import argparse
from datetime import datetime, date
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

TALENT_DB_PATH = r"C:\Users\ADmin2\Desktop\BT_Talent_Database.xlsx"
BREWERY_IG = "bridgeandtunnelbrewery"


def init_talent_db():
    """Create talent database if it doesn't exist."""
    if Path(TALENT_DB_PATH).exists():
        return load_workbook(TALENT_DB_PATH)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Talent Database"
    
    headers = [
        "Band/Act Name", "Genre", "Instagram Handle", "Instagram Followers",
        "Following", "Posts", "Engagement Rate", "Last Updated",
        "Contact Name", "Contact Phone", "Contact Email",
        "Hometown/City", "Draw at B&T", "Times Played B&T",
        "Last Played B&T", "Bring Back?", "Notes"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="C9A84C", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['Q'].width = 40
    
    wb.save(TALENT_DB_PATH)
    print(f"Created new talent database at {TALENT_DB_PATH}")
    return wb


def get_existing_handles(wb):
    """Get set of handles already in database."""
    ws = wb.active
    handles = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            handles.add(str(row[2]).lower().strip())
    return handles


def lookup_socialblade(page, handle):
    """Look up Instagram stats on Social Blade."""
    handle = handle.strip().lstrip("@").lower()
    print(f"  Looking up @{handle} on Social Blade...")
    
    try:
        url = f"https://socialblade.com/instagram/user/{handle}"
        page.goto(url, timeout=30000)
        page.wait_for_timeout(3000)
        
        # Extract followers
        followers = None
        following = None
        posts = None
        
        try:
            # Social Blade follower count
            follower_el = page.locator('#youtube-stats-header-followers-value, [class*="tooltip-2"], .mediaStat').first
            text = page.content()
            
            # Parse from page content
            import re
            
            # Look for followers count
            followers_match = re.search(r'Followers.*?<span[^>]*>([\d,KM]+)</span>', text, re.DOTALL)
            if followers_match:
                followers = followers_match.group(1)
            
            # Try alternative - look for the stats block
            stats_matches = re.findall(r'<p[^>]*style="[^"]*font-size:30px[^"]*"[^>]*>([\d,\.KM]+)</p>', text)
            if stats_matches and len(stats_matches) >= 3:
                posts = stats_matches[0]
                following = stats_matches[1]
                followers = stats_matches[2]
                
        except Exception as e:
            pass
        
        # If Social Blade blocked us, try scraping Instagram directly
        if not followers:
            print(f"    Social Blade blocked — trying Instagram directly...")
            followers, following, posts = scrape_instagram_direct(page, handle)
        
        return {
            "handle": handle,
            "followers": followers or "unknown",
            "following": following or "unknown", 
            "posts": posts or "unknown",
            "found": followers is not None
        }
        
    except Exception as e:
        print(f"    Error looking up @{handle}: {e}")
        return {"handle": handle, "followers": "error", "following": "unknown", "posts": "unknown", "found": False}


def scrape_instagram_direct(page, handle):
    """Scrape Instagram profile directly as fallback."""
    try:
        page.goto(f"https://www.instagram.com/{handle}/", timeout=30000)
        page.wait_for_timeout(4000)
        
        import re
        content = page.content()
        
        # Look for follower count in meta tags or page content
        followers_match = re.search(r'"edge_followed_by":\{"count":(\d+)\}', content)
        if followers_match:
            return format_number(int(followers_match.group(1))), None, None
            
        # Try meta description
        meta_match = re.search(r'([\d,]+)\s+Followers', content)
        if meta_match:
            return meta_match.group(1), None, None
            
        return None, None, None
    except:
        return None, None, None


def format_number(n):
    """Format large numbers nicely."""
    if n >= 1000000:
        return f"{n/1000000:.1f}M"
    elif n >= 1000:
        return f"{n/1000:.1f}K"
    return str(n)


def save_to_database(wb, results):
    """Save research results to talent database."""
    ws = wb.active
    
    # Get existing handles to avoid duplicates
    existing = get_existing_handles(wb)
    
    added = 0
    updated = 0
    
    for result in results:
        handle = result["handle"].lower()
        today = date.today().strftime("%m/%d/%Y")
        
        # Check if handle already exists
        existing_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[2] and str(row[2]).lower().strip() == handle:
                existing_row = row_idx
                break
        
        if existing_row:
            # Update existing row
            ws.cell(row=existing_row, column=4, value=result["followers"])
            ws.cell(row=existing_row, column=5, value=result["following"])
            ws.cell(row=existing_row, column=6, value=result["posts"])
            ws.cell(row=existing_row, column=8, value=today)
            updated += 1
        else:
            # Add new row
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=result.get("name", ""))
            ws.cell(row=next_row, column=2, value=result.get("genre", ""))
            ws.cell(row=next_row, column=3, value=f"@{handle}")
            ws.cell(row=next_row, column=4, value=result["followers"])
            ws.cell(row=next_row, column=5, value=result["following"])
            ws.cell(row=next_row, column=6, value=result["posts"])
            ws.cell(row=next_row, column=8, value=today)
            added += 1
    
    wb.save(TALENT_DB_PATH)
    print(f"\nDatabase updated: {added} added, {updated} updated")


def scrape_brewery_posts(page):
    """Scrape B&T brewery Instagram for show history."""
    print(f"\nScraping @{BREWERY_IG} Instagram for show history...")
    print("Note: Instagram requires login for full scraping.")
    print("Attempting public profile scrape...")
    
    try:
        page.goto(f"https://www.instagram.com/{BREWERY_IG}/", timeout=30000)
        page.wait_for_timeout(5000)
        
        content = page.content()
        
        import re
        
        # Extract post count
        posts_match = re.search(r'"edge_owner_to_timeline_media":\{"count":(\d+)', content)
        followers_match = re.search(r'"edge_followed_by":\{"count":(\d+)\}', content)
        
        if followers_match:
            followers = int(followers_match.group(1))
            print(f"  @{BREWERY_IG} has {format_number(followers)} followers")
        
        # Try to get post URLs
        post_urls = re.findall(r'"shortcode":"([^"]+)"', content)
        print(f"  Found {len(post_urls)} post references on public page")
        
        if not post_urls:
            print("\n  Instagram is blocking public scraping.")
            print("  To scrape brewery post history, Lena needs to:")
            print("  1. Log into Instagram in the browser")
            print("  2. Run this script again with --scrape-brewery --login")
            return []
        
        return post_urls
        
    except Exception as e:
        print(f"  Error scraping brewery: {e}")
        return []


def research_bands(handles, name_map=None):
    """Main function to research a list of Instagram handles."""
    wb = init_talent_db()
    results = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        
        # Set a realistic user agent
        page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        
        for i, handle in enumerate(handles):
            handle = handle.strip().lstrip("@")
            if not handle:
                continue
                
            print(f"\n[{i+1}/{len(handles)}] Researching @{handle}...")
            result = lookup_socialblade(page, handle)
            
            if name_map and handle.lower() in name_map:
                result["name"] = name_map[handle.lower()]
            
            results.append(result)
            print(f"  Followers: {result['followers']}")
            
            # Be polite - don't hammer the server
            if i < len(handles) - 1:
                time.sleep(2)
        
        browser.close()
    
    save_to_database(wb, results)
    
    # Print summary
    print("\n" + "="*50)
    print("RESEARCH SUMMARY")
    print("="*50)
    for r in results:
        status = "✓" if r["found"] else "?"
        print(f"  {status} @{r['handle']}: {r['followers']} followers")
    
    print(f"\nResults saved to: {TALENT_DB_PATH}")
    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Instagram research tool for B&T Brewery")
    parser.add_argument("--handle", help="Instagram handle(s) to research, comma-separated")
    parser.add_argument("--file", help="Text file with one handle per line")
    parser.add_argument("--scrape-brewery", action="store_true", help="Scrape brewery Instagram for show history")
    parser.add_argument("--name", help="Band name for the handle (optional)")
    args = parser.parse_args()
    
    if args.scrape_brewery:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            scrape_brewery_posts(page)
            browser.close()
    elif args.handle:
        handles = [h.strip() for h in args.handle.split(",")]
        name_map = {}
        if args.name and len(handles) == 1:
            name_map[handles[0].lower().lstrip("@")] = args.name
        research_bands(handles, name_map)
    elif args.file:
        with open(args.file, "r") as f:
            handles = [line.strip() for line in f if line.strip()]
        research_bands(handles)
    else:
        print("Usage:")
        print("  python instagram_research.py --handle bandname")
        print("  python instagram_research.py --handle band1,band2,band3")
        print("  python instagram_research.py --scrape-brewery")
        print("  python instagram_research.py --file bands.txt")
