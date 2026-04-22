"""
Bridge and Tunnel Brewery - GitHub Events Page Updater
Fixed: ADmin2 paths, duplicate description, multiple events same date, verification step, graceful git commit
"""
import os, sys, base64, subprocess
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

CALENDAR_PATH = r"C:\Users\ADmin2\Desktop\BT_Events_Calendar.xlsx"
FLYERS_FOLDER = r"C:\Users\ADmin2\Desktop\Instructions\Flyers"
REPO_PATH = r"C:\Users\ADmin2\Desktop\bt-events"
INDEX_PATH = os.path.join(REPO_PATH, "index.html")

def load_events():
    print("Reading events calendar...")
    wb = load_workbook(CALENDAR_PATH)
    ws = wb.active
    events = []
    today = date.today()
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0] and not row[1]:
            continue
        if not row[0]:
            continue
        raw_date = row[0]
        venue = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        flyer_fn = str(row[6] or "").strip()
        ticket = str(row[7] or "").strip()
        event_date = None
        try:
            if isinstance(raw_date, datetime):
                event_date = raw_date.date()
            elif isinstance(raw_date, date):
                event_date = raw_date
            elif isinstance(raw_date, str):
                for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"]:
                    try:
                        event_date = datetime.strptime(raw_date.strip(), fmt).date()
                        break
                    except:
                        continue
        except:
            continue
        if not event_date:
            continue
        if event_date < today:
            continue
        if not flyer_fn:
            continue
        flyer_path = Path(FLYERS_FOLDER) / flyer_fn
        if not flyer_path.exists():
            print(f"  Skipping '{desc or flyer_fn}' - flyer not found: {flyer_fn}")
            continue
        date_display = event_date.strftime("%A, %B %d, %Y").replace(" 0", " ")
        events.append({
            "date": event_date,
            "date_display": date_display,
            "venue": venue,
            "desc": desc,
            "flyer_path": str(flyer_path),
            "ticket": ticket,
        })
    events.sort(key=lambda x: x["date"])
    print(f"Found {len(events)} upcoming events with flyers")
    return events

def build_event_card(event):
    flyer_path = Path(event["flyer_path"])
    ext = flyer_path.suffix.lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png"}
    mime = mime_map.get(ext, "image/jpeg")
    if ext == ".pdf":
        img_html = '<div class="flyer-img" style="background:#2a2a4a;display:flex;align-items:center;justify-content:center;color:#C9A84C;font-size:11px;">PDF Flyer</div>'
    else:
        with open(flyer_path, "rb") as f:
            img_data = base64.b64encode(f.read()).decode()
        img_html = f'<img class="flyer-img" src="data:{mime};base64,{img_data}" alt="Event flyer">'
    ticket = event["ticket"].strip() if event["ticket"] else ""
    desc_lower = event["desc"].lower()
    if ticket.startswith("http"):
        btn = f'<a href="{ticket}" target="_blank" class="event-btn btn-tickets">GET TICKETS</a>'
    elif "free" in desc_lower:
        btn = '<span class="event-btn btn-free">FREE ADMISSION</span>'
    else:
        btn = '<span class="event-btn btn-door">ADMISSION AT DOOR</span>'
    desc_html = f'<div class="event-description">{event["desc"]}</div>' if event["desc"] else ""
    return f'''<div class="event-card">
  <div class="card-top">
    <div class="card-flyer">{img_html}</div>
    <div class="card-middle">
      <div class="event-date">{event["date_display"]}</div>
      {desc_html}
    </div>
    <div class="card-btn-desktop">{btn}</div>
  </div>
  <div class="card-btn-mobile">{btn}</div>
</div>'''

def build_page_html(events):
    ridgewood = [e for e in events if "ridgewood" in e["venue"].lower()]
    liberty = [e for e in events if "liberty" in e["venue"].lower()]
    no_events = '<div class="no-events">No upcoming events — check back soon!</div>'
    ridgewood_html = "\n".join([build_event_card(e) for e in ridgewood]) if ridgewood else no_events
    liberty_html = "\n".join([build_event_card(e) for e in liberty]) if liberty else no_events
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Events — Bridge and Tunnel Brewery</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ background: #C8960C; font-family: Arial, sans-serif; }}
    .loc-nav {{ position: sticky; top: 0; z-index: 100; background: #1A1A2E; display: flex; justify-content: center; gap: 16px; padding: 10px 20px; flex-wrap: wrap; }}
    .loc-nav a {{ color: #C9A84C; text-decoration: none; font-size: 12px; font-weight: bold; letter-spacing: 1px; text-transform: uppercase; padding: 6px 14px; border: 1px solid #C9A84C; border-radius: 4px; }}
    .loc-nav a:hover {{ background: #C9A84C; color: #1A1A2E; }}
    .content {{ padding: 20px; }}
    .venue-section {{ max-width: 960px; margin: 0 auto 50px auto; }}
    .venue-header {{ background: #1A1A2E; text-align: center; padding: 16px 24px; border-radius: 6px; margin-bottom: 20px; }}
    .venue-header h2 {{ color: #C9A84C !important; font-size: 20px; letter-spacing: 2px; text-transform: uppercase; margin: 0; }}
    .venue-header p {{ color: #F5E6C8 !important; font-size: 12px; margin-top: 4px; }}
    .event-card {{ background: #1A1A2E; border-radius: 8px; margin-bottom: 16px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.25); }}
    .card-top {{ display: flex; align-items: stretch; }}
    .card-flyer {{ flex: 0 0 160px; }}
    .flyer-img {{ width: 160px; height: 160px; object-fit: contain; display: block; background: #111; }}
    .card-middle {{ flex: 1; padding: 14px 18px; min-width: 0; }}
    .event-date {{ color: #C9A84C; font-size: 14px; font-weight: bold; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 8px; }}
    .event-description {{ color: #cccccc; font-size: 13px; line-height: 1.7; white-space: pre-wrap; }}
    .card-btn-desktop {{ flex: 0 0 auto; display: flex; align-items: center; padding: 16px; }}
    .card-btn-mobile {{ display: none; }}
    .event-btn {{ display: inline-block; padding: 10px 18px; border-radius: 4px; font-size: 12px; font-weight: bold; text-decoration: none; letter-spacing: 1px; text-align: center; cursor: pointer; white-space: nowrap; }}
    .btn-tickets {{ background: #8B0000; color: #ffffff; }}
    .btn-free {{ background: #2E7D32; color: #ffffff; }}
    .btn-door {{ background: #333333; color: #C9A84C; }}
    .no-events {{ text-align: center; color: #555; font-style: italic; padding: 24px; border: 2px dashed #ccc; border-radius: 8px; background: #fff8dc; }}
    .footer {{ text-align: center; color: #1A1A2E; font-size: 12px; padding: 20px; }}
    .footer a {{ color: #1A1A2E; }}
    @media (max-width: 650px) {{
      .card-btn-desktop {{ display: none; }}
      .card-btn-mobile {{ display: block; padding: 12px 16px; background: transparent; }}
      .card-btn-mobile .event-btn {{ display: block; width: 100%; text-align: center; padding: 12px; font-size: 13px; }}
      .card-flyer {{ flex: 0 0 45%; }}
      .flyer-img {{ width: 100%; height: 100%; min-height: 140px; object-fit: contain;
        background: #111; }}
      .card-middle {{ flex: 1; padding: 10px 12px; }}
      .event-date {{ font-size: 11px; }}
      .event-description {{ font-size: 12px; }}
    }}
  </style>
</head>
<body>
  <nav class="loc-nav">
    <a href="#ridgewood">Ridgewood, Queens</a>
    <a href="#liberty">Liberty, NY</a>
    <a href="https://www.bridgeandtunnelbrewery.com" target="_parent">Home</a>
  </nav>
  <div class="content">
    <div class="venue-section" id="ridgewood">
      <div class="venue-header">
        <h2>Ridgewood, Queens</h2>
        <p>1535 Decatur Street, Ridgewood, NY 11385</p>
      </div>
      {ridgewood_html}
    </div>
    <div class="venue-section" id="liberty">
      <div class="venue-header">
        <h2>Liberty, NY — Catskills</h2>
        <p>50 Lake Street, Liberty, NY 12754</p>
      </div>
      {liberty_html}
    </div>
  </div>
  <div class="footer">
    <p>Bridge and Tunnel Brewery &middot; <a href="https://www.bridgeandtunnelbrewery.com" target="_parent">bridgeandtunnelbrewery.com</a></p>
  </div>
</body>
</html>'''

def verify_html(html, events):
    print("\nVerifying generated HTML...")
    missing = []
    for event in events:
        if event["date_display"] not in html:
            missing.append(f"  MISSING: {event['date_display']} - {event['desc'][:40]}")
    if missing:
        print("WARNING - Some events may not have rendered:")
        for m in missing:
            print(m)
    else:
        print(f"  All {len(events)} events verified in HTML.")
    return len(missing) == 0

def push_to_github():
    print("Pushing to GitHub...")
    os.chdir(REPO_PATH)
    subprocess.run(["git", "add", "."], check=True)
    result = subprocess.run(["git", "commit", "-m", f"Update events {date.today()}"], capture_output=True)
    if result.returncode == 1:
        print("  Nothing new to commit - site already up to date.")
        return
    elif result.returncode != 0:
        raise Exception(f"Git commit failed: {result.stderr.decode()}")
    subprocess.run(["git", "push"], check=True)
    print(f"\nDone! Live at: https://richbigdreamer.github.io/bt-events")

if __name__ == "__main__":
    print("\nBridge and Tunnel Brewery - Events Page Updater")
    print("=" * 50)
    events = load_events()
    if not events:
        print("\nNo upcoming events with flyers found.")
        sys.exit(0)
    print(f"\nBuilding page with {len(events)} event(s)...")
    html = build_page_html(events)
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print("index.html written.")
    verify_html(html, events)
    push_to_github()
